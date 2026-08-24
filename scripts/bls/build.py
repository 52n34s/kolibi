#!/usr/bin/env python3
"""Parse BLS 4.0 Excel into a local foods import CSV and SQL.

Does not write to Supabase. Empty protein/fat/carbs become 0 (NOT NULL columns).
fiber/sugar/sodium currently also coerce to 0 in this pipeline (legacy).
Micronutrients stay blank/NULL when missing (never coerced to 0).

    python3 scripts/bls/download.py
    python3 scripts/bls/build.py
"""

from __future__ import annotations

import csv
import json
import random
import re
import sys
from collections import defaultdict
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(ROOT / "scripts" / "usda_fdc"))
sys.path.insert(0, str(Path(__file__).resolve().parent))

from download import download_all  # noqa: E402
from kolibi_fold import kolibi_fold, kolibi_slug  # noqa: E402

DATA_DIR = ROOT / "data"
IMPORT_CSV_PATH = DATA_DIR / "bls_import.csv"
SKIPPED_PATH = DATA_DIR / "bls_skipped_no_kcal.csv"
COLLISIONS_PATH = DATA_DIR / "bls_slug_collisions.csv"

SOURCE = "bls_4_0"
SOURCE_DESC = "Bundeslebensmittelschlüssel 4.0, Max Rubner-Institut"
SHEET_NAME = "BLS_4_0_Daten_2025_DE"

COL_CODE = "BLS Code"
COL_DE = "Lebensmittelbezeichnung"
COL_EN = "Food name"
COL_KCAL = "ENERCC Energie (Kilokalorien) [kcal/100g]"
COL_PROTEIN = "PROT625 Protein (Nx6,25) [g/100g]"
COL_FAT = "FAT Fett [g/100g]"
COL_CARBS = "CHO Kohlenhydrate, verfügbar [g/100g]"
COL_FIBER = "FIBT Ballaststoffe, gesamt [g/100g]"
COL_SUGAR = "SUGAR Zucker (Mono- und Disaccharide), gesamt [g/100g]"
COL_SODIUM = "NA Natrium [mg/100g]"

# Optional micronutrients / lipids — missing → blank CSV → SQL NULL (never 0).
MICRO_COLUMN_MAP: dict[str, str] = {
    "FE Eisen [mg/100g]": "iron_mg_per_100g",
    "CA Calcium [mg/100g]": "calcium_mg_per_100g",
    "ZN Zink [mg/100g]": "zinc_mg_per_100g",
    "MG Magnesium [mg/100g]": "magnesium_mg_per_100g",
    "K Kalium [mg/100g]": "potassium_mg_per_100g",
    "ID Iodid [µg/100g]": "iodine_ug_per_100g",
    "VITB12 Vitamin B12 (Cobalamine) [µg/100g]": "vitamin_b12_ug_per_100g",
    "FOL Folat-Äquivalent [µg/100g]": "folate_ug_per_100g",
    "VITD Vitamin D [µg/100g]": "vitamin_d_ug_per_100g",
    "FASAT Fettsäuren, gesättigt, gesamt [g/100g]": "saturated_fat_per_100g",
    "FAMS Fettsäure, einfach ungesättigt, gesamt [g/100g]": "monounsaturated_fat_per_100g",
    "FAPU Fettsäuren, mehrfach ungesättigt, gesamt [g/100g]": "polyunsaturated_fat_per_100g",
    "F18:3CN3 Fettsäure C18:3 n-3 all-cis (Alpha-Linolensäure) [g/100g]": "omega3_ala_per_100g",
    "CHORL Cholesterin [mg/100g]": "cholesterol_mg_per_100g",
}
MICRO_COLUMNS = list(MICRO_COLUMN_MAP.values())

REQUIRED_COLUMNS = (
    COL_CODE,
    COL_DE,
    COL_EN,
    COL_KCAL,
    COL_PROTEIN,
    COL_FAT,
    COL_CARBS,
    COL_FIBER,
    COL_SUGAR,
    COL_SODIUM,
    *MICRO_COLUMN_MAP.keys(),
)

CSV_COLUMNS = [
    "name",
    "name_normalized",
    "source",
    "source_ref",
    "kcal_per_100g",
    "protein_per_100g",
    "carbs_per_100g",
    "fat_per_100g",
    "fiber_per_100g",
    "sugar_per_100g",
    "sodium_mg_per_100g",
    *MICRO_COLUMNS,
    "is_countable",
    "category",
    "bls_group",
    "is_verified",
    "names",
    "search_terms",
    "slug",
    "source_desc",
]

TRANSLATED_CSV_PATH = DATA_DIR / "bls_import_translated.csv"
IMPORT_SQL_PATH = DATA_DIR / "bls_import.sql"
SEED_SQL = ROOT / "supabase" / "migrations" / "0008_seed_foods_v3.sql"
USDA_TRANSLATED_CSV = DATA_DIR / "foods_import_usda_fdc_translated.csv"
SQL_BATCH_SIZE = 500

NUTRIENT_MERGE_COLUMNS = [
    "kcal_per_100g",
    "protein_per_100g",
    "carbs_per_100g",
    "fat_per_100g",
    "fiber_per_100g",
    "sugar_per_100g",
    "sodium_mg_per_100g",
    *MICRO_COLUMNS,
]

SQL_COLUMNS = [
    "name",
    "name_normalized",
    "names",
    "search_terms",
    "kcal_per_100g",
    "protein_per_100g",
    "fat_per_100g",
    "carbs_per_100g",
    "fiber_per_100g",
    "sugar_per_100g",
    "sodium_mg_per_100g",
    *MICRO_COLUMNS,
    "is_countable",
    "category",
    "source",
    "source_ref",
    "source_desc",
    "usda_ndb",
    "slug",
    "is_verified",
    "created_by",
]
SQL_UPDATE_COLS = [c for c in SQL_COLUMNS if c not in ("source", "source_ref")]

SEED_REFS = {
    "M713100",
    "M713300",
    "M304600",
    "M602600",
    "M501600",
    "M241300",
    "B221000",
    "B251000",
    "B511000",
    "D771600",
    "C532700",
    "C512300",
    "C515400",
    "K130492",
    "X634012",
    "W140000",
    "W327000",
    "W211200",
    "W211100",
    "W222100",
    "Y036610",
    "F203100",
    "H510802",
    "H520800",
    "Q550000",
    "S570000",
    "S145000",
    "S830000",
    "N330000",
    "F110600",
    "N410100",
    "Y921062",
    "Y943032",
    "X1A2010",
}


def parse_amount(raw: object) -> float | None:
    if raw is None:
        return None
    if isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        value = float(raw)
        if value != value:  # NaN
            return None
        return value
    text = str(raw).strip()
    if not text:
        return None
    try:
        value = float(text.replace(",", "."))
    except ValueError:
        return None
    if value != value:
        return None
    return value


def format_number(value: float | None) -> str:
    if value is None:
        return ""
    if value == int(value):
        return str(int(value))
    return f"{value:.4f}".rstrip("0").rstrip(".")


def format_zero(value: float | None) -> str:
    """Missing protein/fat/carbs (NOT NULL) → 0, never a blank that becomes SQL NULL."""
    return format_number(0.0 if value is None else value)


def cell_text(raw: object) -> str:
    if raw is None:
        return ""
    return str(raw).strip()


def load_rows(xlsx_path: Path) -> list[dict[str, object]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl is not installed (pip install openpyxl)") from None

    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise SystemExit(
                f"missing sheet {SHEET_NAME!r} in {xlsx_path} "
                f"(have {workbook.sheetnames})"
            )
        sheet = workbook[SHEET_NAME]
        iterator = sheet.iter_rows(values_only=True)
        header_row = next(iterator, None)
        if not header_row:
            raise SystemExit(f"empty workbook: {xlsx_path}")
        headers = [cell_text(col) for col in header_row]
        missing = [name for name in REQUIRED_COLUMNS if name not in headers]
        if missing:
            raise SystemExit(f"missing columns: {missing}")
        index = {name: headers.index(name) for name in REQUIRED_COLUMNS}
        rows: list[dict[str, object]] = []
        for raw in iterator:
            rows.append({name: raw[pos] if pos < len(raw) else None for name, pos in index.items()})
        return rows
    finally:
        workbook.close()


def assign_slugs(records: list[dict]) -> list[dict]:
    collisions: list[dict] = []
    grouped: dict[str, list[dict]] = defaultdict(list)
    for record in records:
        base = kolibi_slug(record["name"]) or f"food_{record['source_ref']}"
        record["_base_slug"] = base
        grouped[base].append(record)

    for base, items in grouped.items():
        if len(items) == 1:
            items[0]["slug"] = base
            continue
        for item in items:
            slug = f"{base}_{item['source_ref']}"
            item["slug"] = slug
            collisions.append(
                {
                    "base_slug": base,
                    "source_ref": item["source_ref"],
                    "name": item["name"],
                    "name_de": item["name_de"],
                    "final_slug": slug,
                }
            )
    return collisions


def to_import_row(record: dict) -> dict[str, str]:
    names = {"en": record["name"], "de": record["name_de"], "es": None}
    return {
        "name": record["name"],
        "name_normalized": kolibi_fold(record["name"]),
        "source": SOURCE,
        "source_ref": record["source_ref"],
        "kcal_per_100g": format_number(record["kcal"]),
        "protein_per_100g": format_zero(record["protein"]),
        "carbs_per_100g": format_zero(record["carbs"]),
        "fat_per_100g": format_zero(record["fat"]),
        "fiber_per_100g": format_zero(record["fiber"]),
        "sugar_per_100g": format_zero(record["sugar"]),
        "sodium_mg_per_100g": format_zero(record["sodium_mg"]),
        # Micros: blank when unknown → SQL NULL (never 0).
        **{col: format_number(record.get(col)) for col in MICRO_COLUMNS},
        "is_countable": "false",
        "category": "",
        "bls_group": record["bls_group"],
        "is_verified": "true",
        "names": json.dumps(names, ensure_ascii=False, separators=(",", ":")),
        "search_terms": "[]",
        "slug": record["slug"],
        "source_desc": SOURCE_DESC,
    }


def sql_str(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_text(value: str | None) -> str:
    if value is None:
        return "NULL"
    text = value.strip() if isinstance(value, str) else str(value)
    if text == "":
        return "NULL"
    return sql_str(text)


def sql_num(raw: str | None, zero_if_empty: bool = False) -> str:
    text = (raw or "").strip()
    if not text:
        return "0" if zero_if_empty else "NULL"
    float(text)
    return text


def parse_names(raw: str) -> dict[str, str]:
    try:
        parsed = json.loads(raw or "{}")
    except json.JSONDecodeError:
        parsed = {}
    if not isinstance(parsed, dict):
        parsed = {}
    out: dict[str, str] = {}
    for key in ("en", "de", "es"):
        val = parsed.get(key)
        out[key] = val.strip() if isinstance(val, str) else ""
    return out


def build_search_terms(names: dict[str, str]) -> list[str]:
    seen: set[str] = set()
    terms: list[str] = []
    for key in ("en", "de", "es"):
        folded = kolibi_fold(names.get(key) or "")
        words = folded.split()
        short_food = bool(words) and all(len(w) < 3 for w in words)
        for word in words:
            if len(word) < 3 and not short_food:
                continue
            if word in seen:
                continue
            seen.add(word)
            terms.append(word)
    return terms


def sql_text_array(terms: list[str]) -> str:
    if not terms:
        return "ARRAY[]::text[]"
    inner = ", ".join(sql_str(t) for t in terms)
    return f"ARRAY[{inner}]::text[]"


def sql_names(names: dict[str, str]) -> str:
    payload = {"en": names["en"] or None, "de": names["de"] or None, "es": names["es"] or None}
    dumped = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return sql_str(dumped) + "::jsonb"


def load_seed_slug_owners() -> dict[str, tuple[str, str]]:
    text = SEED_SQL.read_text(encoding="utf-8")
    blocks = re.findall(
        r"VALUES \(\s*'([^']+)',\s*'((?:[^']|'')*)',.*?,\s*'(usda_sr28|bls_4_0)',\s*'([^']+)'",
        text,
        flags=re.S,
    )
    owners: dict[str, tuple[str, str]] = {}
    for slug, _name, source, ref in blocks:
        owners[slug] = (source, ref)
    return owners


def load_usda_slugs() -> set[str]:
    if not USDA_TRANSLATED_CSV.is_file():
        return set()
    rows = list(csv.DictReader(USDA_TRANSLATED_CSV.open(encoding="utf-8")))
    slugs = {row["slug"] for row in rows if row.get("slug")}
    for old, new in (
        ("dates_medjool", "dates_medjool_168191"),
        ("honey", "honey_169640"),
        ("tempeh", "tempeh_174272"),
    ):
        slugs.discard(old)
        slugs.add(new)
    return slugs


def assign_sql_slugs(
    records: list[dict], seed_owners: dict[str, tuple[str, str]], usda_slugs: set[str]
) -> dict:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for rec in records:
        base = kolibi_slug(rec["name"]) or f"food_{rec['source_ref']}"
        rec["base_slug"] = base
        grouped[base].append(rec)

    internal_rows = 0
    internal_groups = 0
    for base, items in grouped.items():
        if len(items) == 1:
            items[0]["slug"] = base
            continue
        internal_groups += 1
        internal_rows += len(items)
        for item in items:
            item["slug"] = f"{base}_{item['source_ref']}"

    vs_seed = []
    vs_usda = []
    used: set[str] = set()
    for rec in records:
        slug = rec["slug"]
        owner = seed_owners.get(slug)
        own_seed = owner == ("bls_4_0", rec["source_ref"])
        collide_seed = owner is not None and not own_seed
        collide_usda = slug in usda_slugs
        if collide_seed or collide_usda:
            new_slug = f"{rec['base_slug']}_{rec['source_ref']}"
            if collide_seed:
                vs_seed.append((rec["source_ref"], rec["name"], slug, owner, new_slug))
            if collide_usda:
                vs_usda.append((rec["source_ref"], rec["name"], slug, new_slug))
            slug = new_slug
            rec["slug"] = slug
        if slug in used:
            raise SystemExit(f"duplicate BLS slug after fallback: {slug}")
        if slug in seed_owners or slug in usda_slugs:
            if not own_seed and slug not in {x[4] for x in vs_seed} and slug not in {
                x[3] for x in vs_usda
            }:
                if slug in seed_owners or slug in usda_slugs:
                    raise SystemExit(f"fallback slug still reserved: {slug}")
        used.add(slug)

    return {
        "internal_rows": internal_rows,
        "internal_groups": internal_groups,
        "vs_seed": vs_seed,
        "vs_usda": vs_usda,
        "own_seed_kept": sum(
            1
            for rec in records
            if seed_owners.get(rec["slug"]) == ("bls_4_0", rec["source_ref"])
        ),
    }


def merge_nutrients_into_translated() -> int:
    if not TRANSLATED_CSV_PATH.is_file():
        raise FileNotFoundError(f"missing translated CSV: {TRANSLATED_CSV_PATH}")
    fresh_rows = list(csv.DictReader(IMPORT_CSV_PATH.open(encoding="utf-8", newline="")))
    by_ref = {row["source_ref"]: row for row in fresh_rows if row.get("source_ref")}
    translated = list(csv.DictReader(TRANSLATED_CSV_PATH.open(encoding="utf-8", newline="")))
    if not translated:
        raise SystemExit("translated CSV is empty")
    fieldnames = list(translated[0].keys())
    for col in NUTRIENT_MERGE_COLUMNS:
        if col not in fieldnames:
            if "sodium_mg_per_100g" in fieldnames and col in MICRO_COLUMNS:
                idx = fieldnames.index("sodium_mg_per_100g") + 1
                while idx < len(fieldnames) and fieldnames[idx] in MICRO_COLUMNS:
                    idx += 1
                fieldnames.insert(idx, col)
            else:
                fieldnames.append(col)
    missing = 0
    for row in translated:
        src = by_ref.get(row.get("source_ref") or "")
        if not src:
            missing += 1
            continue
        for col in NUTRIENT_MERGE_COLUMNS:
            row[col] = src.get(col, "")
    if missing:
        raise SystemExit(f"translated rows missing from fresh CSV: {missing}")
    with TRANSLATED_CSV_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(translated)
    return len(translated)


def tuple_sql(rec: dict) -> str:
    names = rec["names"]
    terms = rec["terms"]
    row = rec["row"]
    values = {
        "name": sql_text(rec["name"]),
        "name_normalized": sql_text(kolibi_fold(rec["name"])),
        "names": sql_names(names),
        "search_terms": sql_text_array(terms),
        "kcal_per_100g": sql_num(row.get("kcal_per_100g")),
        "protein_per_100g": sql_num(row.get("protein_per_100g"), zero_if_empty=True),
        "fat_per_100g": sql_num(row.get("fat_per_100g"), zero_if_empty=True),
        "carbs_per_100g": sql_num(row.get("carbs_per_100g"), zero_if_empty=True),
        "fiber_per_100g": sql_num(row.get("fiber_per_100g"), zero_if_empty=True),
        "sugar_per_100g": sql_num(row.get("sugar_per_100g"), zero_if_empty=True),
        "sodium_mg_per_100g": sql_num(row.get("sodium_mg_per_100g"), zero_if_empty=True),
        **{col: sql_num(row.get(col), zero_if_empty=False) for col in MICRO_COLUMNS},
        "is_countable": "false",
        "category": "NULL",
        "source": sql_text(SOURCE),
        "source_ref": sql_text(rec["source_ref"]),
        "source_desc": sql_text(SOURCE_DESC),
        "usda_ndb": "NULL",
        "slug": sql_text(rec["slug"]),
        "is_verified": "true",
        "created_by": "NULL",
    }
    inner = ",\n    ".join(values[c] for c in SQL_COLUMNS)
    return "  (\n    " + inner + "\n  )"


def chunks(items: list, size: int):
    for i in range(0, len(items), size):
        yield i // size + 1, items[i : i + size]


def write_import_sql() -> dict:
    rows = list(csv.DictReader(TRANSLATED_CSV_PATH.open(encoding="utf-8", newline="")))
    if len(rows) != 7140:
        raise SystemExit(f"expected 7140 CSV rows, got {len(rows)}")

    seed_owners = load_seed_slug_owners()
    usda_slugs = load_usda_slugs()

    records: list[dict] = []
    missing_names = 0
    missing_kcal = 0
    found_seed_refs: set[str] = set()

    for row in rows:
        names = parse_names(row.get("names") or "")
        name = names["en"] or (row.get("name") or "").strip()
        names["en"] = name
        if not names["en"] or not names["de"] or not names["es"]:
            missing_names += 1
        terms = build_search_terms(names)
        if not (row.get("kcal_per_100g") or "").strip():
            missing_kcal += 1
        source_ref = (row.get("source_ref") or "").strip()
        if source_ref in SEED_REFS:
            found_seed_refs.add(source_ref)
        records.append(
            {
                "row": row,
                "names": names,
                "name": name,
                "terms": terms,
                "source_ref": source_ref,
            }
        )

    if found_seed_refs != SEED_REFS:
        raise SystemExit(f"seed refs missing from CSV: {SEED_REFS - found_seed_refs}")
    if missing_kcal:
        raise SystemExit(f"missing kcal: {missing_kcal}")
    if missing_names:
        raise SystemExit(f"missing names.de/en/es: {missing_names}")

    slug_stats = assign_sql_slugs(records, seed_owners, usda_slugs)
    n_batches = (len(records) + SQL_BATCH_SIZE - 1) // SQL_BATCH_SIZE
    col_sql = ",\n  ".join(SQL_COLUMNS)
    update_sql = ",\n  ".join(f"{c} = EXCLUDED.{c}" for c in SQL_UPDATE_COLS)

    header = f"""-- BLS 4.0 import (Bundeslebensmittelschlüssel, Max Rubner-Institut).
-- Generated from data/bls_import_translated.csv
-- Rows: {len(records)} · Batches: {n_batches} × {SQL_BATCH_SIZE} (last batch may be smaller)
-- search_terms: unique kolibi_fold tokens from names.en/de/es;
--   tokens shorter than 3 chars dropped unless the whole name has no longer token.
-- category is NULL; usda_ndb is NULL; is_countable=false; is_verified=true; created_by=NULL.
-- micronutrients: blank CSV → SQL NULL (never 0).
-- ON CONFLICT matches foods_source_ref_unique (partial unique index).
-- The 34 curated bls_4_0 seed rows are updated in place.
-- Run via scripts/bls/run_import.py. This generator does not connect to the DB.

"""
    insert_head = f"""INSERT INTO public.foods (
  {col_sql}
) VALUES
"""
    conflict = f"""
ON CONFLICT (source, source_ref) WHERE source_ref IS NOT NULL
DO UPDATE SET
  {update_sql};
"""

    with IMPORT_SQL_PATH.open("w", encoding="utf-8", newline="\n") as out:
        out.write(header)
        for batch_id, batch in chunks(records, SQL_BATCH_SIZE):
            start = (batch_id - 1) * SQL_BATCH_SIZE + 1
            end = (batch_id - 1) * SQL_BATCH_SIZE + len(batch)
            out.write(f"-- batch {batch_id}/{n_batches} rows {start}-{end}\n")
            out.write(insert_head)
            out.write(",\n".join(tuple_sql(rec) for rec in batch))
            out.write(conflict)
            out.write("\n")
        out.write("NOTIFY pgrst, 'reload schema';\n")

    return {
        "sql_rows": len(records),
        "sql_batches": n_batches,
        "sql_bytes": IMPORT_SQL_PATH.stat().st_size,
        "slug_vs_seed": len(slug_stats["vs_seed"]),
        "slug_vs_usda": len(slug_stats["vs_usda"]),
    }


def build() -> dict:
    xlsx_path = download_all()
    DATA_DIR.mkdir(parents=True, exist_ok=True)

    kept: list[dict] = []
    skipped: list[dict] = []
    seen_codes: set[str] = set()

    for raw in load_rows(xlsx_path):
        source_ref = cell_text(raw[COL_CODE])
        name_de = cell_text(raw[COL_DE])
        name_en = cell_text(raw[COL_EN])
        kcal = parse_amount(raw[COL_KCAL])
        if not source_ref:
            skipped.append(
                {
                    "source_ref": "",
                    "name_de": name_de,
                    "name_en": name_en,
                    "reason": "missing_bls_code",
                    "kcal": "" if kcal is None else format_number(kcal),
                }
            )
            continue
        if source_ref in seen_codes:
            skipped.append(
                {
                    "source_ref": source_ref,
                    "name_de": name_de,
                    "name_en": name_en,
                    "reason": "duplicate_bls_code",
                    "kcal": "" if kcal is None else format_number(kcal),
                }
            )
            continue
        seen_codes.add(source_ref)
        if kcal is None:
            skipped.append(
                {
                    "source_ref": source_ref,
                    "name_de": name_de,
                    "name_en": name_en,
                    "reason": "missing_kcal",
                    "kcal": "",
                }
            )
            continue
        display_name = name_en or name_de
        micros = {
            col: parse_amount(raw[header]) for header, col in MICRO_COLUMN_MAP.items()
        }
        kept.append(
            {
                "source_ref": source_ref,
                "name": display_name,
                "name_de": name_de,
                "bls_group": source_ref[0],
                "kcal": kcal,
                "protein": parse_amount(raw[COL_PROTEIN]),
                "fat": parse_amount(raw[COL_FAT]),
                "carbs": parse_amount(raw[COL_CARBS]),
                "fiber": parse_amount(raw[COL_FIBER]),
                "sugar": parse_amount(raw[COL_SUGAR]),
                "sodium_mg": parse_amount(raw[COL_SODIUM]),
                **micros,
            }
        )

    kept.sort(key=lambda row: row["source_ref"])
    collisions = assign_slugs(kept)

    with SKIPPED_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["source_ref", "name_de", "name_en", "reason", "kcal"],
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(skipped)

    with COLLISIONS_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["base_slug", "source_ref", "name", "name_de", "final_slug"],
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(collisions)

    with IMPORT_CSV_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=CSV_COLUMNS, lineterminator="\n")
        writer.writeheader()
        for record in kept:
            writer.writerow(to_import_row(record))

    translated_rows = merge_nutrients_into_translated()
    sql_stats = write_import_sql()

    return {
        "kept": len(kept),
        "skipped": len(skipped),
        "skipped_missing_kcal": sum(1 for row in skipped if row["reason"] == "missing_kcal"),
        "slug_collision_rows": len(collisions),
        "slug_collision_groups": len({row["base_slug"] for row in collisions}),
        "csv_bytes": IMPORT_CSV_PATH.stat().st_size,
        "translated_rows": translated_rows,
        **sql_stats,
        "samples": random.sample(kept, min(10, len(kept))),
    }


def print_report(stats: dict) -> None:
    print(f"kept: {stats['kept']}")
    print(f"skipped: {stats['skipped']} (missing_kcal={stats['skipped_missing_kcal']})")
    print(
        f"slug collisions: {stats['slug_collision_rows']} rows "
        f"in {stats['slug_collision_groups']} groups"
    )
    print(f"csv: {IMPORT_CSV_PATH} ({stats['csv_bytes']} bytes)")
    print(f"translated: {TRANSLATED_CSV_PATH} ({stats['translated_rows']} rows)")
    print(
        f"sql: {IMPORT_SQL_PATH} ({stats['sql_bytes']} bytes, "
        f"{stats['sql_rows']} rows, {stats['sql_batches']} batches)"
    )
    print("10 random rows:")
    for record in stats["samples"]:
        print(
            f"  {record['slug']} | {record['name']!r} | de={record['name_de']!r} | "
            f"kcal={record['kcal']} group={record['bls_group']} "
            f"ref={record['source_ref']} iron={record.get('iron_mg_per_100g')}"
        )


if __name__ == "__main__":
    try:
        print_report(build())
    except Exception as exc:
        print(f"build failed: {exc}", file=sys.stderr)
        raise
